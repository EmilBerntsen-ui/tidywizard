"""Tests for the Prometheus Panta melting scan and data table parsers."""

from __future__ import annotations

import io

import pandas as pd
import pytest

from core.instruments.prometheus_panta import (
    _classify_measurement_type,
    _flatten_headers,
    _parse_column,
    _pair_columns,
    _split_text_number_unit,
    load_data_table,
    load_melting_scan,
)


# ── _classify_measurement_type ───────────────────────────────────────────


def test_classify_temp() -> None:
    assert _classify_measurement_type("temperatur for cap.1 (\u00b0c)") == "temperature"


def test_classify_ratio() -> None:
    assert _classify_measurement_type("ratio 350 nm / 330 nm for cap.1") == "ratio"


def test_classify_turbidity_turbid() -> None:
    assert _classify_measurement_type("turbidity for cap.1") == "turbidity"


def test_classify_turbidity_scatter() -> None:
    assert _classify_measurement_type("scattering intensity for cap.1") == "turbidity"


def test_classify_cumulant_radius() -> None:
    assert _classify_measurement_type("cumulant radius for cap.1 (nm)") == "cumulant_radius"


def test_classify_unknown_returns_none() -> None:
    assert _classify_measurement_type("unknown signal for cap.1") is None


# ── _parse_column ─────────────────────────────────────────────────────────


def test_parse_column_temperature() -> None:
    cap, mtype = _parse_column("Temperatur for Cap.1 (\u00b0C)")
    assert cap == 1
    assert mtype == "temperature"


def test_parse_column_ratio() -> None:
    cap, mtype = _parse_column("Ratio 350 nm / 330 nm for Cap.3")
    assert cap == 3
    assert mtype == "ratio"


def test_parse_column_turbidity() -> None:
    cap, mtype = _parse_column("Turbidity for Cap.12")
    assert cap == 12
    assert mtype == "turbidity"


def test_parse_column_cumulant_radius() -> None:
    cap, mtype = _parse_column("Cumulant Radius for Cap.2 (nm)")
    assert cap == 2
    assert mtype == "cumulant_radius"


def test_parse_column_variant_capillary_label() -> None:
    """Different firmware versions use different capillary label formats."""
    cap, _ = _parse_column("Ratio 350/330 for Capillary 5")
    assert cap == 5
    cap, _ = _parse_column("Turbidity for Cap 7")
    assert cap == 7


def test_parse_column_no_capillary_raises() -> None:
    with pytest.raises(ValueError, match="capillary number"):
        _parse_column("Ratio 350 nm / 330 nm (no capillary here)")


def test_parse_column_unknown_type_raises() -> None:
    with pytest.raises(ValueError, match="Unrecognised column type"):
        _parse_column("Mystery Signal for Cap.1")


# ── _pair_columns ─────────────────────────────────────────────────────────


def _make_raw_df(n_capillaries: int = 1, n_rows: int = 2) -> pd.DataFrame:
    """Build a minimal raw DataFrame mimicking what pd.read_csv returns."""
    data: dict[str, list] = {}
    for cap in range(1, n_capillaries + 1):
        data[f"Temperatur for Cap.{cap} (C)"] = [f"{25.0 + i * 0.1:.1f}" for i in range(n_rows)]
        data[f"Ratio 350 nm / 330 nm for Cap.{cap}"] = [f"{1.2 + i * 0.01:.3f}" for i in range(n_rows)]
        data[f"Temperatur for Cap.{cap} (C)_t2"] = [f"{25.0 + i * 0.1:.1f}" for i in range(n_rows)]
        data[f"Turbidity for Cap.{cap}"] = [f"{0.002 + i * 0.001:.3f}" for i in range(n_rows)]
        data[f"Temperatur for Cap.{cap} (C)_t3"] = [f"{25.0 + i * 0.1:.1f}" for i in range(n_rows)]
        data[f"Cumulant Radius for Cap.{cap} (nm)"] = [f"{4.5 + i * 0.01:.2f}" for i in range(n_rows)]
    return pd.DataFrame(data)


def test_pair_columns_schema() -> None:
    df_raw = _make_raw_df(n_capillaries=1, n_rows=2)
    out = _pair_columns(df_raw)
    assert list(out.columns) == ["capillary", "measurement_type", "temperature", "value"]


def test_pair_columns_row_count() -> None:
    # 1 capillary × 3 measurement types × 2 rows = 6 rows
    df_raw = _make_raw_df(n_capillaries=1, n_rows=2)
    out = _pair_columns(df_raw)
    assert len(out) == 6


def test_pair_columns_measurement_types() -> None:
    df_raw = _make_raw_df(n_capillaries=1, n_rows=2)
    out = _pair_columns(df_raw)
    assert set(out["measurement_type"].unique()) == {"ratio", "turbidity", "cumulant_radius"}


def test_pair_columns_odd_column_count_raises() -> None:
    df_raw = pd.DataFrame({"Temperatur for Cap.1 (C)": ["25.0"]})
    with pytest.raises(ValueError, match="even number of columns"):
        _pair_columns(df_raw)


# ── Synthetic CSV builder ─────────────────────────────────────────────────


def _make_csv(n_capillaries: int = 2, n_rows: int = 3) -> bytes:
    """
    Build a minimal synthetic Prometheus Panta melting scan CSV (cp1252 encoded).
    Column names use the degree symbol (\\xb0) to test encoding handling.
    """
    header_parts = []
    for cap in range(1, n_capillaries + 1):
        header_parts += [
            f"Temperatur for Cap.{cap} (\xb0C)",
            f"Ratio 350 nm / 330 nm for Cap.{cap}",
            f"Temperatur for Cap.{cap} (\xb0C)",
            f"Turbidity for Cap.{cap}",
            f"Temperatur for Cap.{cap} (\xb0C)",
            f"Cumulant Radius for Cap.{cap} (nm)",
        ]

    rows = [";".join(header_parts)]
    for row_i in range(n_rows):
        temp = 25.0 + row_i * 0.1
        row_parts = []
        for cap in range(1, n_capillaries + 1):
            row_parts += [
                f"{temp:.1f}", f"{1.2 + cap * 0.01 + row_i * 0.001:.3f}",
                f"{temp:.1f}", f"{0.002 + row_i * 0.001:.3f}",
                f"{temp:.1f}", f"{4.5 + cap * 0.1 + row_i * 0.01:.2f}",
            ]
        rows.append(";".join(row_parts))

    return "\n".join(rows).encode("cp1252")


# ── load_melting_scan ─────────────────────────────────────────────────────


def test_load_schema() -> None:
    df = load_melting_scan(_make_csv())
    assert list(df.columns) == ["capillary", "measurement_type", "temperature", "value"]
    assert pd.api.types.is_integer_dtype(df["capillary"])
    assert df["measurement_type"].dtype == object
    assert pd.api.types.is_float_dtype(df["temperature"])
    assert pd.api.types.is_float_dtype(df["value"])


def test_load_row_count() -> None:
    # 2 capillaries × 3 measurement types × 3 rows = 18 rows
    df = load_melting_scan(_make_csv(n_capillaries=2, n_rows=3))
    assert len(df) == 18


def test_load_capillaries_present() -> None:
    df = load_melting_scan(_make_csv(n_capillaries=2))
    assert set(df["capillary"].unique()) == {1, 2}


def test_load_measurement_types_present() -> None:
    df = load_melting_scan(_make_csv())
    assert set(df["measurement_type"].unique()) == {"ratio", "turbidity", "cumulant_radius"}


def test_load_no_duplicate_index() -> None:
    df = load_melting_scan(_make_csv())
    dupes = df.duplicated(subset=["capillary", "measurement_type", "temperature"])
    assert not dupes.any(), "Duplicate (capillary, measurement_type, temperature) rows found"


def test_load_no_nulls_in_output() -> None:
    df = load_melting_scan(_make_csv())
    assert not df.isnull().any().any()


def test_load_deterministic() -> None:
    data = _make_csv()
    a = load_melting_scan(data)
    b = load_melting_scan(data)
    pd.testing.assert_frame_equal(a, b)


def test_load_accepts_bytes() -> None:
    df = load_melting_scan(_make_csv())
    assert len(df) > 0


def test_load_accepts_bytesio() -> None:
    df = load_melting_scan(io.BytesIO(_make_csv()))
    assert len(df) > 0


def test_load_empty_bytes_raises() -> None:
    with pytest.raises(ValueError, match="empty"):
        load_melting_scan(b"")


def test_load_none_raises() -> None:
    with pytest.raises(ValueError, match="empty"):
        load_melting_scan(None)


# ── _flatten_headers ──────────────────────────────────────────────────────


def _make_header_df() -> pd.DataFrame:
    """
    Build a 3-row header DataFrame mimicking a Panta data table export.

    Row 0: section groups (merged cells appear as NaN after the first cell)
    Row 1: sub-headers (only under "Ratio")
    Row 2: leaf column names
    """
    return pd.DataFrame([
        # Row 0: section groups
        [None,      None,     "General",     None,         "Ratio",        None],
        # Row 1: sub-headers
        [None,      None,     None,          None,         "IP # 1 (C)",   None],
        # Row 2: leaf column names
        ["Exclude", "symbol", "Capillaries", "data file",  "ø",            "sigma"],
    ])


def test_flatten_headers_standalone_columns() -> None:
    """Columns with no section header use only the leaf name."""
    cols = _flatten_headers(_make_header_df())
    assert cols[0] == "Exclude"
    assert cols[1] == "symbol"


def test_flatten_headers_two_level() -> None:
    """Section + leaf name are joined with '_'."""
    cols = _flatten_headers(_make_header_df())
    assert cols[2] == "General_Capillaries"
    assert cols[3] == "General_data file"


def test_flatten_headers_three_level() -> None:
    """Section + subsection + leaf name are joined with '_'."""
    cols = _flatten_headers(_make_header_df())
    assert cols[4] == "Ratio_IP # 1 (C)_ø"
    assert cols[5] == "Ratio_IP # 1 (C)_sigma"


def test_flatten_headers_no_leaf_name() -> None:
    """A column with only a section group (no leaf) uses just the section name."""
    header = pd.DataFrame([
        ["General", None],
        [None,      None],
        [None,      None],   # leaf row intentionally empty
    ])
    cols = _flatten_headers(header)
    assert cols[0] == "General"


def test_flatten_headers_duplicate_names_get_suffix() -> None:
    """Duplicate column names get .1, .2, … suffixes to stay unique."""
    header = pd.DataFrame([
        ["General", "General"],
        [None,      None],
        [None,      None],
    ])
    cols = _flatten_headers(header)
    assert cols[0] == "General"
    assert cols[1] == "General.1"


def test_flatten_headers_leaf_row_not_forward_filled() -> None:
    """Values in the leaf row must NOT bleed into adjacent empty leaf cells."""
    header = pd.DataFrame([
        ["General", "General", "General"],
        [None,      None,      None],
        ["Cap",     None,      None],   # only first leaf cell is named
    ])
    cols = _flatten_headers(header)
    # The second and third columns should NOT inherit "Cap" from the first
    assert cols[1] == "General"
    assert cols[2] == "General.1"


# ── load_data_table ───────────────────────────────────────────────────────


def _make_data_table_xlsx(n_data_rows: int = 2) -> bytes:
    """
    Build a minimal synthetic Panta data table as an in-memory .xlsx file.

    Header structure (3 rows):
        Row 0:  [NaN,     NaN,      General,      General,    Ratio,    Ratio ]
        Row 1:  [NaN,     NaN,      NaN,          NaN,        IP#1,     IP#1  ]
        Row 2:  [Exclude, symbol,   Capillaries,  data file,  ø,        sigma ]
    Data rows: Exclude=False, symbol=X, Capillaries=1..N, etc.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header rows (merged cells represented by only setting the first cell)
    ws.append([None,      None,     "General",    "General",  "Ratio",  "Ratio"])
    ws.append([None,      None,     None,         None,       "IP#1",   "IP#1"])
    ws.append(["Exclude", "symbol", "Capillaries","data file","ø",      "sigma"])

    for i in range(1, n_data_rows + 1):
        ws.append([False, "X", i, f"scan_{i}.csv", 65.0 + i, 0.1])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_load_data_table_schema() -> None:
    df = load_data_table(io.BytesIO(_make_data_table_xlsx()))
    assert "Exclude" in df.columns
    assert "General_Capillaries" in df.columns
    assert "Ratio_IP#1_ø" in df.columns


def test_load_data_table_row_count() -> None:
    df = load_data_table(io.BytesIO(_make_data_table_xlsx(n_data_rows=5)))
    assert len(df) == 5


def test_load_data_table_no_data_rows_returns_empty() -> None:
    """A file with only header rows returns an empty DataFrame (not an error)."""
    df = load_data_table(io.BytesIO(_make_data_table_xlsx(n_data_rows=0)))
    assert len(df) == 0
    assert "Exclude" in df.columns


def test_load_data_table_none_raises() -> None:
    with pytest.raises(ValueError, match="empty"):
        load_data_table(None)


def test_load_data_table_empty_bytes_raises() -> None:
    with pytest.raises(ValueError, match="empty"):
        load_data_table(b"")


# ── _split_text_number_unit ───────────────────────────────────────────────


def _make_components_df(values: list) -> pd.DataFrame:
    return pd.DataFrame({"Viscosity_components": values})


def test_split_basic() -> None:
    df = _split_text_number_unit(_make_components_df(["Sodium acetate 25 mM"]), "Viscosity_components")
    assert "Viscosity_components_name" in df.columns
    assert "Viscosity_components_value_mM" in df.columns
    assert df["Viscosity_components_name"].iloc[0] == "Sodium acetate"
    assert df["Viscosity_components_value_mM"].iloc[0] == 25.0


def test_split_trailing_colon_stripped() -> None:
    """Trailing ':' in the name (e.g. 'Sodium acetate: 25 mM') is stripped."""
    df = _split_text_number_unit(_make_components_df(["Sodium acetate: 25 mM"]), "Viscosity_components")
    assert df["Viscosity_components_name"].iloc[0] == "Sodium acetate"


def test_split_float_value() -> None:
    df = _split_text_number_unit(_make_components_df(["HEPES 7.5 mM"]), "Viscosity_components")
    assert df["Viscosity_components_value_mM"].iloc[0] == 7.5


def test_split_original_column_removed() -> None:
    df = _split_text_number_unit(_make_components_df(["NaCl 150 mM"]), "Viscosity_components")
    assert "Viscosity_components" not in df.columns


def test_split_columns_inserted_in_place() -> None:
    """The two new columns appear where the original column was."""
    df = pd.DataFrame({"A": [1], "Viscosity_components": ["NaCl 150 mM"], "B": [2]})
    df = _split_text_number_unit(df, "Viscosity_components")
    assert list(df.columns) == ["A", "Viscosity_components_name", "Viscosity_components_value_mM", "B"]


def test_split_null_row_becomes_na() -> None:
    # No unit can be inferred from a null row, so column is named without unit suffix
    df = _split_text_number_unit(_make_components_df([None]), "Viscosity_components")
    assert pd.isna(df["Viscosity_components_name"].iloc[0])
    assert pd.isna(df["Viscosity_components_value"].iloc[0])


def test_split_no_match_name_kept_value_na() -> None:
    """A value that doesn't match the pattern keeps raw text in name, NA in value."""
    df = _split_text_number_unit(_make_components_df(["PBS"]), "Viscosity_components")
    assert df["Viscosity_components_name"].iloc[0] == "PBS"
    assert pd.isna(df["Viscosity_components_value"].iloc[0])


def test_split_mixed_units_raises() -> None:
    df = _make_components_df(["NaCl 150 mM", "Tris 10 µM"])
    with pytest.raises(ValueError, match="mixed units"):
        _split_text_number_unit(df, "Viscosity_components")


def test_load_data_table_auto_splits_viscosity_components() -> None:
    """load_data_table splits Viscosity_components regardless of capitalisation."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Viscosity", "Viscosity"])
    ws.append([None, None])
    ws.append(["Components", "other"])   # capital C — real Panta files use this
    ws.append(["Sodium acetate 25 mM", "foo"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    df = load_data_table(buf)
    assert "Viscosity_Components" not in df.columns
    assert "Viscosity_Components_name" in df.columns
    assert "Viscosity_Components_value_mM" in df.columns
    assert df["Viscosity_Components_name"].iloc[0] == "Sodium acetate"
    assert df["Viscosity_Components_value_mM"].iloc[0] == 25.0


def test_load_data_table_drops_viscosity_solvent() -> None:
    """load_data_table drops Viscosity_Solvent (any casing) automatically."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Viscosity", "Viscosity"])
    ws.append([None, None])
    ws.append(["Solvent", "other"])
    ws.append(["Sodium Acetate Buffer", "foo"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    df = load_data_table(buf)
    assert "Viscosity_Solvent" not in df.columns
    assert "Viscosity_other" in df.columns
